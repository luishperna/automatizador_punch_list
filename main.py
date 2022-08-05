from tkinter import Frame, Tk, Label, Button, Entry, messagebox, ttk, END, PhotoImage
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.platypus import Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, mm
from datetime import datetime
from openpyxl import load_workbook


# Salva as opções do usuário incremantadas nas listas
qtd_bornes = lista_bornes_de_passagem = []
qtd_capacitor = lista_capacitor_cilindrico = []
qtd_gavetas = lista_gavetas = []
qtd_inversor = lista_inversor_de_frequencia = []
qtd_mca = lista_modulos_mca = []
qtd_soft = lista_soft_starter = []
qtd_switch = lista_switch_gerenciavel = []
qtd_tampa_traseira = lista_tampa_traseira = []
qtd_tampa_lateral = lista_tampa_lateral = []


# Função pega a opção do usuário e trabalha nas condições incrementando nas listas


# Função para atualizar os dados no excel
def atualizar_excel(evento=None):
    try:
        get_se = str(se.get())
        get_ccm = int(ccm.get())
        get_coluna = str(coluna.get()).upper()
        get_gaveta = str(gaveta.get()).lower()
        get_pendencia = str(pendencia.get())
        get_codigo = str(codigo.get()).upper()
        get_especificacao = especificacao.get()
        get_quantidade = int(quantidade.get())
        get_responsabilidade = str(responsabilidade.get()).upper()
        acao = 'Enviar ao cliente final para ser instalado durante o comissionamento'
        data_atual = datetime.now()
        data = data_atual.strftime('%d/%m/%y')

        # SE formatada para buscar entre as abas
        se_e_numero = 'SE' + get_se

        juntando_pendencia_especificacao = get_pendencia + ' ' + get_especificacao
        pendencia_completa = juntando_pendencia_especificacao.upper()


# ----------------------------------------------------------------------------------------------------------
        # Jogando as descrições e quantidades nas listas respectivas para usar na geração do pdf
        # if get_pendencia == opcoes_pendencias[0]:
        #     get_quantidade.append(qtd_bornes)



    except (ValueError):
        messagebox.showerror('Erro: campo vazio',
                             'Ops... um dos campos está vazio\nFavor preencher todos os campos!')

    try:
        # Abrindo arquivo .xlsx
        planilha = load_workbook('Punch list (cliente) SE.xlsx')

        # Se tiver SE irá fazer as alterações
        if se_e_numero in planilha.sheetnames:
            aba_se = planilha[se_e_numero]
            for linha_a, linha_b, linha_c, linha_d, linha_e, linha_f, linha_g, linha_h, linha_i in zip(aba_se['A'], aba_se['B'], aba_se['C'], aba_se['D'], aba_se['E'], aba_se['F'], aba_se['G'], aba_se['H'], aba_se['I']):
                # Se todas as condições forem verdadeiras irá atualizar apenas o valor
                if (linha_a.value != None) and (linha_a.value == get_ccm) and \
                    (linha_b.value != None) and (linha_b.value == get_coluna) and \
                    (linha_c.value != None) and (linha_c.value == get_gaveta) and \
                    (linha_d.value != None) and (linha_d.value == pendencia_completa) and \
                        (linha_e.value != None) and (linha_e.value == get_codigo):
                    quantidade_f = linha_f.value + get_quantidade
                    linha_f.value = quantidade_f
                    break

                # Se uma das condições não for verdadeira, irá criar nova linha para inserir as informações
                elif (linha_a.value != get_ccm) or (linha_b.value != get_coluna) or (linha_c.value != get_gaveta) or (linha_d.value != pendencia_completa) or (linha_e.value != get_codigo):
                    if linha_a.value == None:
                        linha_a.value = get_ccm
                        linha_b.value = get_coluna
                        linha_c.value = get_gaveta
                        linha_d.value = pendencia_completa
                        linha_e.value = get_codigo
                        linha_f.value = get_quantidade
                        linha_g.value = acao
                        linha_h.value = data
                        linha_i.value = get_responsabilidade
                        break

        # Se ainda não tiver SE irá adicionar a nova aba com as informações do primeiro item
        else:
            aba_base = planilha['BASE']
            nova_aba = planilha.copy_worksheet(aba_base)
            nova_aba.title = se_e_numero
            planilha[se_e_numero]['A1'].value = f'Punch list (cliente) SE-{get_se}'
            planilha[se_e_numero]['A3'].value = get_ccm
            planilha[se_e_numero]['B3'].value = get_coluna
            planilha[se_e_numero]['C3'].value = get_gaveta
            planilha[se_e_numero]['D3'].value = pendencia_completa
            planilha[se_e_numero]['E3'].value = get_codigo
            planilha[se_e_numero]['F3'].value = get_quantidade
            planilha[se_e_numero]['G3'].value = acao
            planilha[se_e_numero]['H3'].value = data
            planilha[se_e_numero]['I3'].value = get_responsabilidade


        # Salva as atualizações feitas no arquivo .xlsx
        planilha.save(filename='Punch list (cliente) SE.xlsx')

    except (PermissionError):
        messagebox.showerror('Erro: planilha aberta',
                             'Ops... a planilha está aberta no momento\nFavor fecha-lá antes de atualizar!')
    else:
        # Limpando os campos
        coluna.delete(0, END)
        gaveta.delete(0, END)
        codigo.delete(0, END)
        pendencia.delete(0, END)
        especificacao.delete(0, END)
        quantidade.delete(0, END)


# Função para gerar um arquivo em .pdf
def gerar_pdf():
    messagebox.showinfo("Ops...", "Em desenvolvimento ainda...")
#     try:
#         dados_tabela = [
#             ["Item", "Descrição", "Qtd"],
#             ["1", "Bornes de passagem", str(sum(qtd_bornes))],
#             ["2", "Capacitor cilíndrico", str(sum(qtd_capacitor))],
#             ["3", "Gaveta", str(sum(qtd_gavetas))],
#             ["4", "Inversor de frequência", str(sum(qtd_inversor))],
#             ["5", "Módulo MCA", str(sum(qtd_mca))],
#             ["6", "Soft-starter", str(sum(qtd_soft))],
#             ["7", "Switch gerenciável", str(sum(qtd_switch))],
#             ["8", "Tampa traseira", str(sum(qtd_tampa_traseira))],
#             ["9", "Tampa lateral", str(sum(qtd_tampa_lateral))],
#         ]

#         # Pegando a o pro9jeto digitado para incrementar no arquivo .pdf
#         projeto_escolhido = str(definir_projeto.get()).upper()
#         texto_projeto_escolhido = (f'Projeto: {projeto_escolhido}')

#         # Pegando a data e a hora atual para incrementar no arquivo .pdf
#         data_e_hora_atuais = datetime.now()
#         data_e_hora_em_texto = data_e_hora_atuais.strftime(
#             'Relatório emitido em %d/%m/%Y às %H:%M')

#         # Estilos para os parágrafos
#         titulo_style = ParagraphStyle('Heading1', fontName='Helvetica-Bold',
#                                       fontSize=14, textColor=colors.black, leading=20, alignment=1, spaceAfter=30)

#         paragrafo_projeto_style = ParagraphStyle('Heading2', fontName='Helvetica',
#                                                  fontSize=12, textColor=colors.black, leading=20, alignment=1, spaceAfter=30)

#         paragrafo_style = ParagraphStyle('Heading3', fontName='Helvetica',
#                                          fontSize=12, textColor=colors.black, leading=20, alignment=1, spaceAfter=15)

#         paragrafo_informacoes_style = ParagraphStyle('Heading4', fontName='Helvetica',
#                                                      fontSize=12, textColor=colors.black, leading=20, alignment=1, spaceBefore=242)

#         paragrafo_final_style = ParagraphStyle('Heading5', fontName='Helvetica',
#                                                fontSize=8, textColor=colors.black, leading=20, alignment=1)

#         # Parágrafos do pdf
#         titulo = Paragraph(
#             'Relatório - Relação de Envio Posterior', titulo_style)

#         paragrafo_texto_projeto_escolhido = Paragraph(
#             texto_projeto_escolhido, paragrafo_projeto_style)

#         paragrafo = Paragraph(
#             'Segue abaixo a relação dos itens para envio posterior:', paragrafo_style)

#         paragrafo_informacoes = Paragraph(
#             data_e_hora_em_texto, paragrafo_informacoes_style)

#         paragrafo_final = Paragraph(
#             'Automatizador - Relação de Envio Posterior developed by Luís Henrique Perna © 2022', paragrafo_final_style)

#         # Tabela do pdf já com estilização
#         tabela = Table(dados_tabela, [12*mm, 50*mm, 12*mm], 10*[10*mm])
#         tabela.setStyle(TableStyle([
#             ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
#             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#             ('ALIGN', (1, 1), (1, -1), 'LEFT'),
#             ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
#             ('BOX', (0, 0), (-1, -1), 0.25, None),
#             ('BACKGROUND', (0, 0), (-1, 0), colors.green),
#             ('BACKGROUND', (0, 2), (-1, 2), colors.gainsboro),
#             ('BACKGROUND', (0, 4), (-1, 4), colors.gainsboro),
#             ('BACKGROUND', (0, 6), (-1, 6), colors.gainsboro),
#             ('BACKGROUND', (0, 8), (-1, 8), colors.gainsboro),
#             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#             ('TEXTCOLOR', (0, 0), (-1, 0), colors.white)
#         ]))

#         # Cria o arquivo .pdf e construí a página
#         doc = SimpleDocTemplate("Relatório - Relação de Envio Posterior.pdf", pagesize=A4,
#                                 rightMargin=10*mm, leftMargin=10*mm, topMargin=20*mm)

#         doc.build([titulo, paragrafo_texto_projeto_escolhido, paragrafo, tabela,
#                   paragrafo_informacoes, paragrafo_final])

#         # Alteração nos Label de confirmação ao usuário
#         texto_confirmacao_emitido.config(text="Relatório emitido com sucesso!")

#         texto_retirar_na_pasta.config(
#             text="Para pegá-lo favor acessar a pasta do programa")

#     except (PermissionError):
#         messagebox.showerror("ERRO: PDF Aberto",
#                              "Favor fechar o relatório em pdf aberto para emitir o novo!")


# Cores usadas na interface
cor_azul = "#116BAC"
cor_cinza = "#7D7D7D"
fonte_branca = "#FFFFFF"
cor_de_fundo = "#F8F8FF"


# Opções de pendência para ComboBox da interface
opcoes_pendencias = ["Bornes de passagem",
                     "Capacitor cilíndrico",
                     "Gaveta",
                     "Inversor de frequência",
                     "Módulo MCA",
                     "Soft-starter",
                     "Switch gerenciável",
                     "Tampa traseira",
                     "Tampa lateral"]


# Opções de clientes para ComboBox da interface
opcoes_clientes = ["Cliente Final", "Empresa"]


# Cria a interface
home = Tk()
home.title('Automatizador - Relação de Envio Posterior v1.8')
home.resizable(width=0, height=0)
home.configure(bg=cor_de_fundo)
home.iconbitmap('img\\icone_sincronizar.ico')


# Centralizando a janela da interface na tela
largura_janela = 734
altura_janela = 449

largura_tela = home.winfo_screenwidth()
altura_tela = home.winfo_screenheight()

centro_x = int(largura_tela / 2 - largura_janela / 2)
centro_y = int(altura_tela / 2 - altura_janela / 2)

home.geometry(f'{largura_janela}x{altura_janela}+{centro_x}+{centro_y}')


# Cria um frame na interface para os inputs do usuário
frame_entry = Frame(home)
frame_entry.configure(bg=cor_azul)
frame_entry.grid(column=0, row=4, padx=50, ipady=3)

# Cria um frame na interface para os inputs do usuário
frame_botoes = Frame(home)
frame_botoes.configure(bg=cor_de_fundo)
frame_botoes.grid(column=0, row=6, pady=(0, 20), ipady=3)


# Texto inicial do programa
texto_inicial = Label(home, text='Automatizador de Relação de Envio Posterior')
texto_inicial.configure(bg=cor_azul)
texto_inicial.configure(fg=fonte_branca)
texto_inicial["font"] = ("Verdana", "12", "bold")
texto_inicial.grid(column=0, row=0, ipady=10, sticky='ew')


# Texto para informando para digitar nos campos abaixo
texto_de_orientacao = Label(
    home, text='DIGITE AS INFORMAÇÕES NOS CAMPOS ABAIXO')
texto_de_orientacao.configure(bg=cor_de_fundo)
texto_de_orientacao.configure(fg=cor_cinza)
texto_de_orientacao["font"] = ("Verdana", "10", "bold")
texto_de_orientacao.grid(column=0, row=1, pady=(20, 15))


# Texto SE (dentro do frame)
texto_se = Label(home, text='SE:')
texto_se.configure(bg=cor_de_fundo)
texto_se.configure(fg=cor_cinza)
texto_se["font"] = ("Verdana", "10", "bold")
texto_se.grid(column=0, row=2)


# Cria input para o usuário escolher uma opção (dentro do frame)
se = Entry(home, width=20, borderwidth=3, relief="groove")
se.configure(bg="#F5FFFA")
se.grid(column=0, row=3, pady=(0, 25))
se.focus()


# Texto CCM (dentro do frame)
texto_ccm = Label(frame_entry, text='CCM:')
texto_ccm.configure(bg=cor_azul)
texto_ccm.configure(fg=fonte_branca)
texto_ccm["font"] = ("Verdana", "8", "bold")
texto_ccm.grid(column=0, row=0, padx=(20, 10), pady=8)


# Cria input para o usuário escolher uma opção (dentro do frame)
ccm = Entry(frame_entry, width=20)
ccm.configure(bg="#F5FFFA")
ccm.grid(column=0, row=1, padx=(20, 10))


# Texto COLUNA (dentro do frame)
texto_coluna = Label(frame_entry, text='COLUNA:')
texto_coluna.configure(bg=cor_azul)
texto_coluna.configure(fg=fonte_branca)
texto_coluna["font"] = ("Verdana", "8", "bold")
texto_coluna.grid(column=1, row=0, padx=10, pady=8)


# Cria input para o usuário digitar o CCM (dentro do frame)
coluna = Entry(frame_entry, width=20)
coluna.configure(bg="#F5FFFA")
coluna.grid(column=1, row=1, padx=10)


# Texto GAVETA (dentro do frame)
texto_gaveta = Label(frame_entry, text='GAVETA:')
texto_gaveta.configure(bg=cor_azul)
texto_gaveta.configure(fg=fonte_branca)
texto_gaveta["font"] = ("Verdana", "8", "bold")
texto_gaveta.grid(column=2, row=0, padx=10, pady=8)


# Cria input para o usuário digitar a COLUNA (dentro do frame)
gaveta = Entry(frame_entry, width=20)
gaveta.configure(bg="#F5FFFA")
gaveta.grid(column=2, row=1, padx=10)


# Texto CÓDIGO (dentro do frame)
texto_codigo = Label(frame_entry, text='CÓDIGO:')
texto_codigo.configure(bg=cor_azul)
texto_codigo.configure(fg=fonte_branca)
texto_codigo["font"] = ("Verdana", "8", "bold")
texto_codigo.grid(column=3, row=0, padx=(10, 20), pady=8)


# Cria input para o usuário digitar a GAVETA (dentro do frame)
codigo = Entry(frame_entry, width=20)
codigo.configure(bg="#F5FFFA")
codigo.grid(column=3, row=1, padx=(10, 20))


# Texto PENDÊNCIA (dentro do frame)
texto_pendencia = Label(frame_entry, text='PENDÊNCIA:')
texto_pendencia.configure(bg=cor_azul)
texto_pendencia.configure(fg=fonte_branca)
texto_pendencia["font"] = ("Verdana", "8", "bold")
texto_pendencia.grid(column=0, row=2, padx=(20, 10), pady=(25, 8))


# ComboBox para o usuário escolher a pendência (dentro do frame)
pendencia = ttk.Combobox(frame_entry, values=opcoes_pendencias)
pendencia.grid(column=0, row=3, padx=(20, 10), pady=(0, 8))


# Texto ESPECIFICAÇÃO (dentro do frame)
texto_especificacao = Label(frame_entry, text='ESPECIFICAÇÃO:')
texto_especificacao.configure(bg=cor_azul)
texto_especificacao.configure(fg=fonte_branca)
texto_especificacao["font"] = ("Verdana", "8", "bold")
texto_especificacao.grid(column=1, row=2, padx=10, pady=(25, 8))


# Cria input para o usuário digitar a ESPECIFICAÇÃO (dentro do frame)
especificacao = Entry(frame_entry, width=20)
especificacao.configure(bg="#F5FFFA")
especificacao.grid(column=1, row=3, padx=10, pady=(0, 8))


# Texto QUANTIDADE (dentro do frame)
texto_quantidade = Label(frame_entry, text='QUANTIDADE:')
texto_quantidade.configure(bg=cor_azul)
texto_quantidade.configure(fg=fonte_branca)
texto_quantidade["font"] = ("Verdana", "8", "bold")
texto_quantidade.grid(column=2, row=2, padx=10, pady=(25, 8))


# Cria input para o usuário digitar a QUANTIDADE (dentro do frame)
quantidade = Entry(frame_entry, width=20)
quantidade.configure(bg="#F5FFFA")
quantidade.grid(column=2, row=3, padx=10, pady=(0, 8))


# Texto RESPONSABILIDADE (dentro do frame)
texto_responsabilidade = Label(frame_entry, text='RESPONSABILIDADE:')
texto_responsabilidade.configure(bg=cor_azul)
texto_responsabilidade.configure(fg=fonte_branca)
texto_responsabilidade["font"] = ("Verdana", "8", "bold")
texto_responsabilidade.grid(column=3, row=2, padx=(10, 20), pady=(25, 8))


# ComboBox para o usuário escolher a pendência (dentro do frame)
responsabilidade = ttk.Combobox(frame_entry, values=opcoes_clientes)
responsabilidade.grid(column=3, row=3, padx=(10, 20), pady=(0, 8))


# Exibe texto informativo
texto_orientacao_excel = Label(
    home, text='Clique em ATUALIZAR EXCEL ou na tecla ENTER para enviar as informações')
texto_orientacao_excel.configure(bg=cor_de_fundo)
texto_orientacao_excel.configure(fg=cor_cinza)
texto_orientacao_excel["font"] = ("Verdana", "10", "bold")
texto_orientacao_excel.grid(column=0, row=5, pady=20)


# Botão executa a função input_usuario() (dentro do frame)
botao_excel = Button(frame_botoes, text='ATUALIZAR EXCEL',
                     command=atualizar_excel, borderwidth=3, relief="ridge")
botao_excel.configure(bg=cor_azul)
botao_excel.configure(fg=fonte_branca)
botao_excel["font"] = ("Verdana", "10", "bold")
botao_excel.grid(column=0, row=0, padx=40, ipadx=10)


# Botão executa a função gerar_pdf() (dentro do frame)
botao_relatório = Button(frame_botoes, text='EMITIR RELATÓRIO',
                         command=gerar_pdf, borderwidth=3, relief="ridge")
botao_relatório.configure(bg=cor_azul)
botao_relatório.configure(fg=fonte_branca)
botao_relatório["font"] = ("Verdana", "10", "bold")
botao_relatório.grid(column=1, row=0, padx=40, ipadx=10)


# Citando o desenvolvedor
texto_credito = Label(
    home, text='Automatizador - Relação de Envio Posterior developed by Luís Henrique Perna © 2022')
texto_credito.configure(bg=cor_azul)
texto_credito.configure(fg=fonte_branca)
texto_credito["font"] = ("Verdana", "8")
texto_credito.grid(column=0, row=7, sticky='ew')


# Ao apertar Enter executa a função atualizar_excel()
home.bind('<Return>', atualizar_excel)


# Mantém a interface aberta (loop)
home.mainloop()
